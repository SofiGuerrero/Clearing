import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
HEADER = PatternFill("solid", fgColor="1F497D")

header_font = Font(bold=True, color="FFFFFF", size=11)
data_font   = Font(size=10)
center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left", vertical="center")
thin        = Side(style="thin", color="AAAAAA")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_header(ws, headers, col_widths):
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = HEADER
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def write_pivot_sheet(wb, pivot):
    ws = wb.active
    ws.title = "Pivot Clearing"

    headers    = ["G/L Account", "Company Code", "Balance", "Status"]
    col_widths = [15, 16, 20, 22]
    _write_header(ws, headers, col_widths)

    for row_idx, row in pivot.iterrows():
        excel_row = row_idx + 2
        balance   = row["Balance"]
        status    = "CLEAR" if balance == 0 else "OPEN ITEMS"
        fill      = GREEN if balance == 0 else RED

        values = [row["G/L Account"], row["Company Code"], balance, status]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = data_font
            cell.border    = border
            cell.alignment = center

        ws.cell(row=excel_row, column=3).number_format = '#,##0.00'


def write_docs_sheet(wb, df_filtered, pivot):
    """Hoja con los document numbers de las combinaciones con balance = 0."""
    ws = wb.create_sheet("Docs a Clearear")

    headers    = ["G/L Account", "Company Code", "Document Number", "Amount in LC", "Text", "Assignment"]
    col_widths = [15, 16, 20, 18, 45, 25]
    _write_header(ws, headers, col_widths)

    # Solo combinaciones con balance 0
    zero_pairs = set(
        zip(pivot[pivot["Balance"] == 0]["G/L Account"],
            pivot[pivot["Balance"] == 0]["Company Code"])
    )

    # Filtrar rows que pertenecen a combinaciones con balance 0
    mask = df_filtered.apply(
        lambda r: (r["Account"], r["Company Code"]) in zero_pairs, axis=1
    )
    docs = df_filtered[mask].copy()
    docs = docs.sort_values(["Account", "Company Code", "Document Number"]).reset_index(drop=True)

    excel_row = 2
    for _, row in docs.iterrows():
        values = [
            row["Account"],
            row["Company Code"],
            row["Document Number"],
            row["Amount in Local Currency"],
            row.get("Text", ""),
            row.get("Assignment", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill      = GREEN
            cell.font      = data_font
            cell.border    = border
            cell.alignment = center if col_idx != 5 else left

        ws.cell(row=excel_row, column=4).number_format = '#,##0.00'
        excel_row += 1


def write_resumen_sheet(wb, pivot):
    zero_count = (pivot["Balance"] == 0).sum()
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Total combinaciones"
    ws2["B1"] = len(pivot)
    ws2["A2"] = "Con balance = 0 (CLEAR)"
    ws2["B2"] = int(zero_count)
    ws2["A3"] = "Con balance != 0 (OPEN ITEMS)"
    ws2["B3"] = int(len(pivot) - zero_count)

    for c in ["A1", "A2", "A3"]:
        ws2[c].font = Font(bold=True)
