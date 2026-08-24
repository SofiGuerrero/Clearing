"""
SAP Clearing Dashboard — Servidor standalone (todo en un archivo)
Requiere: pip install flask pywin32 pandas openpyxl
"""

import os
import sys
import queue
import threading
import calendar
import pandas as pd
import openpyxl
import win32com.client
from datetime import date
from flask import Flask, Response, render_template_string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── DIRECTORIO DE SALIDA ─────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "SAP Clearing")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── ESTILOS EXCEL ────────────────────────────────────────────────────────────
GREEN       = PatternFill("solid", fgColor="C6EFCE")
RED         = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="1F497D")
header_font = Font(bold=True, color="FFFFFF", size=11)
data_font   = Font(size=10)
center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left", vertical="center")
thin        = Side(style="thin", color="AAAAAA")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_header(ws, headers, col_widths):
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def write_pivot_sheet(wb, pivot):
    ws = wb.active
    ws.title = "Pivot Clearing"
    _write_header(ws, ["G/L Account", "Company Code", "Balance", "Status"], [15, 16, 20, 22])
    for row_idx, row in pivot.iterrows():
        excel_row = row_idx + 2
        balance   = row["Balance"]
        status    = "CLEAR" if balance == 0 else "OPEN ITEMS"
        fill      = GREEN if balance == 0 else RED
        for col_idx, val in enumerate([row["G/L Account"], row["Company Code"], balance, status], start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = fill; cell.font = data_font
            cell.border = border; cell.alignment = center
        ws.cell(row=excel_row, column=3).number_format = '#,##0.00'


def write_docs_sheet(wb, df_filtered, pivot):
    ws = wb.create_sheet("Docs a Clearear")
    _write_header(ws, ["G/L Account", "Company Code", "Document Number", "Amount in LC", "Text", "Assignment"],
                  [15, 16, 20, 18, 45, 25])
    zero_pairs = set(zip(pivot[pivot["Balance"] == 0]["G/L Account"],
                         pivot[pivot["Balance"] == 0]["Company Code"]))
    mask = df_filtered.apply(lambda r: (r["Account"], r["Company Code"]) in zero_pairs, axis=1)
    docs = df_filtered[mask].copy().sort_values(["Account", "Company Code", "Document Number"]).reset_index(drop=True)
    for excel_row, (_, row) in enumerate(docs.iterrows(), start=2):
        for col_idx, val in enumerate([row["Account"], row["Company Code"], row["Document Number"],
                                       row["Amount in Local Currency"], row.get("Text", ""), row.get("Assignment", "")], start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = GREEN; cell.font = data_font
            cell.border = border; cell.alignment = center if col_idx != 5 else left
        ws.cell(row=excel_row, column=4).number_format = '#,##0.00'


def write_resumen_sheet(wb, pivot):
    zero_count = (pivot["Balance"] == 0).sum()
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Total combinaciones";    ws2["B1"] = len(pivot)
    ws2["A2"] = "Con balance = 0 (CLEAR)"; ws2["B2"] = int(zero_count)
    ws2["A3"] = "Con balance != 0 (OPEN ITEMS)"; ws2["B3"] = int(len(pivot) - zero_count)
    for c in ["A1", "A2", "A3"]:
        ws2[c].font = Font(bold=True)


# ─── SAP HELPERS ──────────────────────────────────────────────────────────────
def get_last_day_of_month():
    today = date.today()
    last = calendar.monthrange(today.year, today.month)[1]
    return date(today.year, today.month, last).strftime("%Y%m%d")


def get_sap_session(system_prefix):
    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
    except Exception:
        raise RuntimeError("No se pudo conectar al SAP GUI. Asegurate de que SAP GUI esté abierto.")
    sap_app = sap_gui_auto.GetScriptingEngine
    for i in range(sap_app.Children.Count):
        conn = sap_app.Children(i)
        if conn.Children.Count > 0 and conn.Children(0).PassportSystemid[:3] == system_prefix:
            return conn.Children(0)
    raise RuntimeError(f"Por favor abrí la conexión {system_prefix} en SAP GUI.")


def export_to_excel(session, primary_coords, output_dir, filename):
    exported = False

    # Modo 1: lista clásica — intentar distintas coordenadas de label
    for coords in [primary_coords, "1,2", "1,1", "2,2", "1,3", "17,2"]:
        try:
            session.findById(f"wnd[0]/usr/lbl[{coords}]").setFocus()
            session.findById("wnd[0]").sendVKey(16)
            session.findById("wnd[1]/tbar[0]/btn[20]").press()
            exported = True
            print(f"  Export modo lista clásica (lbl[{coords}])")
            break
        except Exception:
            pass

    if not exported:
        # Modo 2: ALV Grid — buscar el shell container
        for shell_path in [
            "wnd[0]/usr/cntlGRID1/shellcont/shell",
            "wnd[0]/usr/cntlALV_CONTAINER_1/shellcont/shell",
            "wnd[0]/usr/subSUBSCREEN:SAPLKKBL:0600/cntlGRID1/shellcont/shell",
        ]:
            try:
                shell = session.findById(shell_path)
                shell.pressToolbarButton("&MB_EXPORT")
                try:
                    session.findById("wnd[1]/tbar[0]/btn[20]").press()
                except Exception:
                    pass
                exported = True
                print(f"  Export modo ALV grid ({shell_path})")
                break
            except Exception:
                pass

    if not exported:
        raise RuntimeError(
            "No se pudo iniciar el export. Verificar que FAGLL03 muestre resultados "
            "y que no haya popups abiertos en SAP GUI."
        )

    session.findById("wnd[1]/usr/ctxtDY_PATH").text = output_dir
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = filename
    session.findById("wnd[1]/tbar[0]/btn[11]").press()


# ─── US ───────────────────────────────────────────────────────────────────────
US_INPUT  = os.path.join(OUTPUT_DIR, "Clearing US Input.xlsx")
US_OUTPUT = os.path.join(OUTPUT_DIR, "Clearing US Pivot.xlsx")


def download_us():
    s_last_day = get_last_day_of_month()
    session = get_sap_session("ISP")
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "FAGLL03"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    session.findById("wnd[1]/usr/txtV-LOW").text = "/GL ACC SA"
    session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").setFocus()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").caretPosition = 7
    session.findById("wnd[0]").sendVKey(4)
    shell = session.findById("wnd[1]/usr/cntlCONTAINER/shellcont/shell")
    shell.focusDate = s_last_day
    shell.firstVisibleDate = s_last_day
    shell.selectionInterval = f"{s_last_day},{s_last_day}"
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    export_to_excel(session, "17,14", OUTPUT_DIR, "Clearing US Input.xlsx")
    print(f"Reporte US descargado. Fecha clave: {s_last_day}")


def apply_account_rules_us(df):
    frames = []
    for acc, group in df.groupby("Account"):
        if acc == 172501:
            mask = (group["Text"].str.contains(r"ADP TAX", case=False, na=False) |
                    group["Text"].str.contains(r"CIT EMPLOYEE", case=False, na=False))
            filtered = group[mask]
            excluded = len(group) - len(filtered)
            if excluded:
                print(f"  172501: excluidos {excluded} rows")
        elif acc == 176000:
            mask = (group["Text"].str.contains(r"RECLASS", case=False, na=False) |
                    group["Text"].str.contains(r"WAGES & SALARIES TO BE PAID", case=False, na=False))
            filtered = group[mask]
            excluded = len(group) - len(filtered)
            if excluded:
                print(f"  176000: excluidos {excluded} rows")
        elif acc == 179000:
            def is_valid(val):
                if pd.isna(val) or str(val).strip() == "": return True
                s = str(val).strip()
                return s.upper() == "PAYROLL" or s.replace("0", "").isdigit() or s.isdigit()
            mask = group["Assignment"].apply(is_valid)
            filtered = group[mask]
            excluded = len(group) - len(filtered)
            if excluded:
                print(f"  179000: excluidos {excluded} rows")
        else:
            filtered = group
        frames.append(filtered)
    return pd.concat(frames, ignore_index=True) if frames else df.iloc[0:0]


def pivot_us():
    print(f"Leyendo: {US_INPUT}")
    df = pd.read_excel(US_INPUT)
    print(f"  {len(df):,} registros cargados.")
    before = len(df)
    df = df[df["Year/Month"] >= "2024/01"]
    print(f"  Filtro Year/Month: excluidos {before - len(df):,}. Quedan {len(df):,}.")
    df = apply_account_rules_us(df)
    pivot = df.pivot_table(index=["Account", "Company Code"], values="Amount in Local Currency", aggfunc="sum").reset_index()
    pivot.columns = ["G/L Account", "Company Code", "Balance"]
    pivot["Balance"] = pivot["Balance"].round(2)
    pivot = pivot.sort_values(["G/L Account", "Company Code"]).reset_index(drop=True)
    zero_count = (pivot["Balance"] == 0).sum()
    print(f"  Balance 0: {zero_count} | Total: {len(pivot)}")
    wb = openpyxl.Workbook()
    write_pivot_sheet(wb, pivot); write_docs_sheet(wb, df, pivot); write_resumen_sheet(wb, pivot)
    wb.save(US_OUTPUT)
    print(f"Archivo generado: {US_OUTPUT}")
    print(f"  CLEAR: {zero_count} | OPEN ITEMS: {len(pivot) - zero_count}")


def run_us():
    download_us()
    pivot_us()


# ─── CANADA ───────────────────────────────────────────────────────────────────
CA_INPUT  = os.path.join(OUTPUT_DIR, "Clearing Canada Input.xlsx")
CA_OUTPUT = os.path.join(OUTPUT_DIR, "Clearing Canada Pivot.xlsx")


def download_canada():
    s_last_day = get_last_day_of_month()
    session = get_sap_session("ISP")
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "FAGLL03"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    session.findById("wnd[1]/usr/txtV-LOW").text = "/GLCANADA"
    session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
    session.findById("wnd[1]/usr/txtV-LOW").caretPosition = 9
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").setFocus()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").caretPosition = 7
    session.findById("wnd[0]").sendVKey(4)
    shell = session.findById("wnd[1]/usr/cntlCONTAINER/shellcont/shell")
    shell.focusDate = s_last_day
    shell.firstVisibleDate = s_last_day
    shell.selectionInterval = f"{s_last_day},{s_last_day}"
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    export_to_excel(session, "17,14", OUTPUT_DIR, "Clearing Canada Input.xlsx")
    print(f"Reporte Canada descargado. Fecha clave: {s_last_day}")


def pivot_canada():
    print(f"Leyendo: {CA_INPUT}")
    xl = pd.ExcelFile(CA_INPUT)
    sheet = "Data" if "Data" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(CA_INPUT, sheet_name=sheet)
    print(f"  Hoja: '{sheet}' | {len(df):,} registros cargados.")
    before = len(df)
    df = df[df["Year/Month"] >= "2024/01"]
    print(f"  Filtro Year/Month: excluidos {before - len(df):,}. Quedan {len(df):,}.")
    pivot = df.pivot_table(index=["Account", "Company Code"], values="Amount in Local Currency", aggfunc="sum").reset_index()
    pivot.columns = ["G/L Account", "Company Code", "Balance"]
    pivot["Balance"] = pivot["Balance"].round(2)
    pivot = pivot.sort_values(["G/L Account", "Company Code"]).reset_index(drop=True)
    zero_count = (pivot["Balance"] == 0).sum()
    print(f"  Balance 0: {zero_count} | Total: {len(pivot)}")
    wb = openpyxl.Workbook()
    write_pivot_sheet(wb, pivot); write_docs_sheet(wb, df, pivot); write_resumen_sheet(wb, pivot)
    wb.save(CA_OUTPUT)
    print(f"Archivo generado: {CA_OUTPUT}")
    print(f"  CLEAR: {zero_count} | OPEN ITEMS: {len(pivot) - zero_count}")


def run_canada():
    download_canada()
    pivot_canada()


# ─── CONCUR ───────────────────────────────────────────────────────────────────
CO_INPUT  = os.path.join(OUTPUT_DIR, "Clearing Concur Input.xlsx")
CO_OUTPUT = os.path.join(OUTPUT_DIR, "Clearing Concur Pivot.xlsx")


def download_concur():
    s_last_day = get_last_day_of_month()
    session = get_sap_session("I4P")
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = "FAGLL03"
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    session.findById("wnd[1]/usr/txtV-LOW").text = "PAYROLLGLCLEAR"
    session.findById("wnd[1]/usr/txtENAME-LOW").text = "I751884"
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").setFocus()
    session.findById("wnd[0]/usr/ctxtPA_STIDA").caretPosition = 9
    session.findById("wnd[0]").sendVKey(4)
    shell = session.findById("wnd[1]/usr/cntlCONTAINER/shellcont/shell")
    shell.focusDate = s_last_day
    shell.firstVisibleDate = s_last_day
    shell.selectionInterval = f"{s_last_day},{s_last_day}"
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    export_to_excel(session, "20,11", OUTPUT_DIR, "Clearing Concur Input.xlsx")
    print(f"Reporte Concur descargado. Fecha clave: {s_last_day}")


def pivot_concur():
    print(f"Leyendo: {CO_INPUT}")
    df = pd.read_excel(CO_INPUT, sheet_name="Data")
    print(f"  {len(df):,} registros cargados.")
    before = len(df)
    df = df[df["Clearing Document"].isna()]
    print(f"  Excluidos {before - len(df):,} con clearing document. Quedan {len(df):,}.")
    before = len(df)
    df = df[df["Year/Month"] >= "2024/01"]
    print(f"  Filtro Year/Month: excluidos {before - len(df):,}. Quedan {len(df):,}.")
    pivot = df.pivot_table(index=["Account", "Company Code"], values="Amount in Local Currency", aggfunc="sum").reset_index()
    pivot.columns = ["G/L Account", "Company Code", "Balance"]
    pivot["Balance"] = pivot["Balance"].round(2)
    pivot = pivot.sort_values(["G/L Account", "Company Code"]).reset_index(drop=True)
    zero_count = (pivot["Balance"] == 0).sum()
    print(f"  Balance 0: {zero_count} | Total: {len(pivot)}")
    wb = openpyxl.Workbook()
    write_pivot_sheet(wb, pivot); write_docs_sheet(wb, df, pivot); write_resumen_sheet(wb, pivot)
    wb.save(CO_OUTPUT)
    print(f"Archivo generado: {CO_OUTPUT}")
    print(f"  CLEAR: {zero_count} | OPEN ITEMS: {len(pivot) - zero_count}")


def run_concur():
    download_concur()
    pivot_concur()


# ─── RUNNERS ──────────────────────────────────────────────────────────────────
RUNNERS = {"us": run_us, "canada": run_canada, "concur": run_concur}

_lock = threading.Lock()


class _QueueWriter:
    def __init__(self, q):
        self.q = q
        self._buf = ""

    def write(self, text):
        self._buf += text
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line.strip():
                self.q.put(line)

    def flush(self):
        if self._buf.strip():
            self.q.put(self._buf.strip())
            self._buf = ""


def _run_captured(func, q):
    writer = _QueueWriter(q)
    old_out, old_err = sys.stdout, sys.stderr
    sys.stdout = writer
    sys.stderr = writer
    try:
        func()
        writer.flush()
        q.put("__DONE__:OK")
    except Exception as e:
        writer.flush()
        q.put(f"ERROR: {e}")
        q.put("__DONE__:ERROR")
    finally:
        sys.stdout = old_out
        sys.stderr = old_err


# ─── FLASK ────────────────────────────────────────────────────────────────────
app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SAP Clearing Dashboard</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #f0f2f5; min-height: 100vh; }
  header {
    background: #1a2b4a; color: #fff; padding: 20px 32px;
    display: flex; align-items: center; gap: 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,.3);
  }
  header h1 { font-size: 1.4rem; font-weight: 600; }
  header small { font-size: .8rem; color: #94a3b8; margin-top: 2px; display: block; }
  .grid {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(340px, 1fr));
    gap: 24px; padding: 32px; max-width: 1200px; margin: 0 auto;
  }
  .card { background: #fff; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,.08); overflow: hidden; }
  .card-header { padding: 20px 24px 16px; border-left: 5px solid; display: flex; align-items: center; gap: 12px; }
  .card-header .icon { font-size: 2rem; }
  .card-header h2 { font-size: 1.1rem; font-weight: 600; color: #1a2b4a; }
  .card-header p  { font-size: .82rem; color: #666; margin-top: 2px; }
  .card-us     .card-header { border-color: #2563eb; }
  .card-canada .card-header { border-color: #dc2626; }
  .card-concur .card-header { border-color: #d97706; }
  .card-body { padding: 16px 24px 20px; }
  .btn {
    display: inline-flex; align-items: center; gap: 8px;
    padding: 9px 22px; border: none; border-radius: 6px;
    font-size: .9rem; font-weight: 600; cursor: pointer; transition: opacity .15s;
  }
  .btn:disabled { opacity: .5; cursor: not-allowed; }
  .btn-us     { background: #2563eb; color: #fff; }
  .btn-canada { background: #dc2626; color: #fff; }
  .btn-concur { background: #d97706; color: #fff; }
  .btn:not(:disabled):hover { opacity: .85; }
  .status {
    display: inline-flex; align-items: center; gap: 6px;
    font-size: .8rem; font-weight: 600; padding: 4px 10px;
    border-radius: 20px; margin-left: 10px; vertical-align: middle;
  }
  .status.idle    { background: #e5e7eb; color: #6b7280; }
  .status.running { background: #dbeafe; color: #1d4ed8; }
  .status.ok      { background: #dcfce7; color: #15803d; }
  .status.error   { background: #fee2e2; color: #b91c1c; }
  .spinner {
    width: 12px; height: 12px; border: 2px solid currentColor;
    border-top-color: transparent; border-radius: 50%;
    animation: spin .7s linear infinite; display: inline-block;
  }
  @keyframes spin { to { transform: rotate(360deg); } }
  .terminal {
    margin-top: 14px; background: #0f172a; color: #94a3b8;
    font-family: 'Cascadia Code', 'Consolas', monospace; font-size: .78rem;
    border-radius: 6px; padding: 12px 14px; height: 200px;
    overflow-y: auto; white-space: pre-wrap; word-break: break-all; display: none;
  }
  .terminal.visible { display: block; }
  .line-ok   { color: #4ade80; }
  .line-err  { color: #f87171; }
  .line-info { color: #38bdf8; }
</style>
</head>
<body>
<header>
  <span style="font-size:1.6rem">&#128202;</span>
  <div>
    <h1>SAP Clearing Dashboard</h1>
    <small>Los archivos se guardan en Documentos &rsaquo; SAP Clearing</small>
  </div>
</header>
<div class="grid">
  <div class="card card-us">
    <div class="card-header">
      <div class="icon">&#127482;&#127480;</div>
      <div><h2>ISP &mdash; United States</h2><p>Variante /GL ACC SA &middot; sistema ISP</p></div>
    </div>
    <div class="card-body">
      <button class="btn btn-us" onclick="runScript('us',this)">&#9654; Ejecutar</button>
      <span id="status-us" class="status idle">&#9679; Idle</span>
      <div id="term-us" class="terminal"></div>
    </div>
  </div>
  <div class="card card-canada">
    <div class="card-header">
      <div class="icon">&#127464;&#127462;</div>
      <div><h2>ISP &mdash; Canada</h2><p>Variante /GLCANADA &middot; sistema ISP</p></div>
    </div>
    <div class="card-body">
      <button class="btn btn-canada" onclick="runScript('canada',this)">&#9654; Ejecutar</button>
      <span id="status-canada" class="status idle">&#9679; Idle</span>
      <div id="term-canada" class="terminal"></div>
    </div>
  </div>
  <div class="card card-concur">
    <div class="card-header">
      <div class="icon">&#128203;</div>
      <div><h2>Concur</h2><p>Variante PAYROLLGLCLEAR &middot; sistema I4P</p></div>
    </div>
    <div class="card-body">
      <button class="btn btn-concur" onclick="runScript('concur',this)">&#9654; Ejecutar</button>
      <span id="status-concur" class="status idle">&#9679; Idle</span>
      <div id="term-concur" class="terminal"></div>
    </div>
  </div>
</div>
<script>
function runScript(key, btn) {
  const term = document.getElementById('term-' + key);
  const status = document.getElementById('status-' + key);
  btn.disabled = true;
  term.innerHTML = '';
  term.classList.add('visible');
  status.className = 'status running';
  status.innerHTML = '<span class="spinner"></span> Corriendo...';
  const es = new EventSource('/run/' + key);
  es.onmessage = function(e) {
    const msg = e.data;
    if (msg.startsWith('__DONE__:')) {
      es.close(); btn.disabled = false;
      const ok = msg.includes(':OK');
      status.className = 'status ' + (ok ? 'ok' : 'error');
      status.innerHTML  = ok ? '&#10003; Completado' : '&#10007; Error';
      term.scrollTop = term.scrollHeight; return;
    }
    const line = document.createElement('div');
    const lower = msg.toLowerCase();
    if (lower.includes('error')) line.className = 'line-err';
    else if (lower.includes('correctamente') || lower.includes('completado') || lower.includes('clear:')) line.className = 'line-ok';
    else if (lower.includes('leyendo') || lower.includes('archivo generado') || lower.includes('generado')) line.className = 'line-info';
    line.textContent = msg;
    term.appendChild(line);
    term.scrollTop = term.scrollHeight;
  };
  es.onerror = function() {
    es.close(); btn.disabled = false;
    status.className = 'status error';
    status.innerHTML = '&#10007; Sin conexi&oacute;n con el servidor';
  };
}
</script>
</body>
</html>"""


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/run/<key>")
def run_script(key):
    if key not in RUNNERS:
        return "Script no encontrado", 404

    q = queue.Queue()
    t = threading.Thread(target=_run_captured, args=(RUNNERS[key], q), daemon=True)
    t.start()

    def generate():
        while True:
            try:
                msg = q.get(timeout=600)
                yield f"data: {msg}\n\n"
                if msg.startswith("__DONE__"):
                    break
            except queue.Empty:
                yield "data: ERROR: Timeout (10 min)\n\n"
                yield "data: __DONE__:ERROR\n\n"
                break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


if __name__ == "__main__":
    import webbrowser
    print(f"Archivos de salida en: {OUTPUT_DIR}")
    print("Iniciando SAP Clearing Dashboard en http://localhost:5001")
    webbrowser.open("http://localhost:5001")
    app.run(host="0.0.0.0", port=5001, threaded=True)
